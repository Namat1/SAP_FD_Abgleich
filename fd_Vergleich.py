import io
import re
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Feste Prüfregeln
# ---------------------------------------------------------------------------

DAY_NAMES = {
    1: "Montag",
    2: "Dienstag",
    3: "Mittwoch",
    4: "Donnerstag",
    5: "Freitag",
    6: "Samstag",
}

DAY_SHORT = {
    1: "Mo",
    2: "Die",
    3: "Mitt",
    4: "Don",
    5: "Fr",
    6: "Sam",
}

DAY_COLUMNS_TOUR = {
    1: 6,   # G = Mo
    2: 7,   # H = Die
    3: 8,   # I = Mitt
    4: 9,   # J = Don
    5: 10,  # K = Fr
    6: 11,  # L = Sam
}

DAY_COLUMN_CANDIDATES = {
    1: ["mo", "montag"],
    2: ["die", "di", "dienstag"],
    3: ["mitt", "mit", "mi", "mittwoch"],
    4: ["don", "do", "donnerstag"],
    5: ["fr", "frei", "freitag"],
    6: ["sam", "sa", "samstag"],
}

# Aus Direkt werden nur diese Touren geprüft.
DIRECT_ALLOWED_TOURS: Set[str] = {"1058", "2058", "3058", "4058", "5058", "6030"}

SAP_COL_INDEX = 0       # SAP-Datei Fallback: A = SAP Nummer
SAP_DAY_COL_INDEX = 6   # SAP-Datei Fallback: G = Liefertag
TOUR_CSB_COL_INDEX = 0  # Tourenplanung Fallback: A = CSB
TOUR_SAP_COL_INDEX = 1  # Tourenplanung Fallback: B = SAP


# ---------------------------------------------------------------------------
# Helfer
# ---------------------------------------------------------------------------

def normalize_header_name(value) -> str:
    text = "" if value is None or pd.isna(value) else str(value)
    text = text.strip().lower()
    text = (
        text.replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
    )
    return "".join(ch for ch in text if ch.isalnum())


def value_to_clean_text(value) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_sap_series(series: pd.Series) -> pd.Series:
    if series.empty:
        return series.astype(str)
    numeric = pd.to_numeric(series, errors="coerce")
    is_int = numeric.notna() & (numeric == numeric.round())
    out = series.astype(str)
    out = out.where(~is_int, numeric.where(is_int).astype("Int64").astype(str))
    out = out.str.strip()
    out = out.replace({"nan": "", "<NA>": "", "None": "", "NaT": ""})
    return out


def normalize_day_code_series(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    text = series.astype(str).map(normalize_header_name)
    text_map = {
        "1": 1, "mo": 1, "montag": 1,
        "2": 2, "di": 2, "die": 2, "dienstag": 2,
        "3": 3, "mi": 3, "mitt": 3, "mit": 3, "mittwoch": 3,
        "4": 4, "do": 4, "don": 4, "donnerstag": 4,
        "5": 5, "fr": 5, "frei": 5, "freitag": 5,
        "6": 6, "sa": 6, "sam": 6, "samstag": 6,
    }
    mapped = text.map(text_map)
    return numeric.where(numeric.notna(), mapped)


def pick_first_matching_column(columns: List[str], candidates: List[str]) -> Optional[str]:
    candidate_set = {normalize_header_name(c) for c in candidates}
    for column in columns:
        if normalize_header_name(column) in candidate_set:
            return column
    return None


def pick_column_by_name_or_position(columns: List[str], candidates: List[str], fallback_index: Optional[int]) -> Optional[str]:
    found = pick_first_matching_column(columns, candidates)
    if found is not None:
        return found
    if fallback_index is not None and len(columns) > fallback_index:
        return columns[fallback_index]
    return None


def make_unique_columns(raw_columns: List[object]) -> List[str]:
    result: List[str] = []
    seen: Dict[str, int] = {}
    for index, value in enumerate(raw_columns, start=1):
        name = value_to_clean_text(value)
        if not name:
            name = f"Spalte_{index}"
        count = seen.get(name, 0) + 1
        seen[name] = count
        if count > 1:
            name = f"{name}_{count}"
        result.append(name)
    return result


def read_excel_with_detected_header(excel: pd.ExcelFile, sheet_name: str, kind: str) -> pd.DataFrame:
    """Erkennt die eigentliche Kopfzeile selbst.

    kind='tour': erwartet SAP plus Tages-Spalten Mo bis Sam.
    kind='sap': erwartet SAP plus Liefertag.
    """
    raw = pd.read_excel(excel, sheet_name=sheet_name, header=None, dtype=object)
    if raw.empty:
        return pd.DataFrame()

    day_names_flat = {
        normalize_header_name(candidate)
        for values in DAY_COLUMN_CANDIDATES.values()
        for candidate in values
    }

    header_row: Optional[int] = None
    max_scan_rows = min(len(raw), 30)

    for row_index in range(max_scan_rows):
        values = [normalize_header_name(value) for value in raw.iloc[row_index].tolist()]
        value_set = set(values)
        has_sap = bool(value_set & {"sap", "sapnummer", "sapnr", "sapnum", "kundennummer", "kundennr"})

        if kind == "tour":
            day_hits = sum(1 for value in values if value in day_names_flat)
            if has_sap and day_hits >= 2:
                header_row = row_index
                break
        else:
            has_day = bool(value_set & {"liefertag", "liefertagcode", "liefercode", "lt", "tag"})
            if has_sap and has_day:
                header_row = row_index
                break

    if header_row is None:
        return pd.read_excel(excel, sheet_name=sheet_name, header=0, dtype=object)

    df = raw.iloc[header_row + 1:].copy()
    df.columns = make_unique_columns(raw.iloc[header_row].tolist())
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def day_value_is_set(value) -> bool:
    if value is None or pd.isna(value):
        return False
    text = str(value).strip()
    if not text:
        return False
    if text.lower() in {"nan", "none", "<na>", "-", "--"}:
        return False
    number = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.notna(number) and float(number) == 0:
        return False
    return True


def extract_tour_numbers(value) -> List[str]:
    """Liest Tournummern aus Zellwerten wie 1058, 1058 / 2058, 22+3."""
    if value is None or pd.isna(value):
        return []
    text = value_to_clean_text(value)
    if not text:
        return []
    return re.findall(r"\d+", text)


def select_relevant_sheets(excel: pd.ExcelFile) -> Dict[str, str]:
    """Gibt die zu prüfenden Blätter zurück: Direkt, NMS, Malchow. MK wird nicht geprüft."""
    result: Dict[str, str] = {}
    for sheet_name in excel.sheet_names:
        norm = normalize_header_name(sheet_name)
        if "direkt" == norm or norm.endswith("direkt") or "direkt" in norm:
            result.setdefault("Direkt", sheet_name)
        elif "nms" in norm or "neumuenster" in norm or "neumunster" in norm:
            result.setdefault("NMS", sheet_name)
        elif "malchow" in norm:
            result.setdefault("Malchow", sheet_name)
    return result


def get_tour_columns(df: pd.DataFrame) -> Dict[int, str]:
    columns = list(df.columns)
    out: Dict[int, str] = {}
    for day_num, fallback_index in DAY_COLUMNS_TOUR.items():
        col = pick_column_by_name_or_position(columns, DAY_COLUMN_CANDIDATES[day_num], fallback_index)
        if col is not None:
            out[day_num] = col
    return out


def sorted_day_text(days: Set[int]) -> str:
    if not days:
        return "(keine hinterlegt)"
    return ", ".join(f"{d} {DAY_NAMES[d]}" for d in sorted(days))


def _tour_field_has_allowed_direct(value) -> bool:
    """Prüft ob mindestens eine Tournummer im Feld zu den freigegebenen gehört."""
    nums = set(re.findall(r"\d+", value_to_clean_text(value)))
    return bool(nums & DIRECT_ALLOWED_TOURS)


def apply_final_scope_filter(tour_df: pd.DataFrame) -> pd.DataFrame:
    """Letzte Sicherheitsprüfung für den Prüfbereich.

    NMS und Malchow bleiben vollständig enthalten.
    Direkt bleibt ausschließlich enthalten, wenn die konkrete Tournummer
    eine der freigegebenen Direkt-Touren ist. Dadurch kann niemals das
    komplette Direkt-Blatt in den Vergleich laufen.
    """
    if tour_df.empty or "Bereich" not in tour_df.columns:
        return tour_df

    work = tour_df.copy()
    if "Tournummer Tour" in work.columns:
        work["Tournummer Tour"] = work["Tournummer Tour"].map(value_to_clean_text)

    direct_mask = work["Bereich"].eq("Direkt")
    # Harter Check: exakter Match ODER mindestens eine Nummer im Feld erlaubt
    direct_ok = direct_mask & (
        work["Tournummer Tour"].isin(DIRECT_ALLOWED_TOURS)
        | work["Tournummer Tour"].apply(_tour_field_has_allowed_direct)
    )
    other_ok = work["Bereich"].isin({"NMS", "Malchow"})

    filtered = work[other_ok | direct_ok].copy()
    return filtered.reset_index(drop=True)


def merge_info(target: Dict[str, Dict[str, str]], sap: str, info: Dict[str, str]) -> None:
    existing = target.setdefault(sap, {"csb": "", "name": "", "strasse": "", "plz": "", "ort": "", "bereich": "", "blatt": ""})
    for key, value in info.items():
        if value and not existing.get(key):
            existing[key] = value


# ---------------------------------------------------------------------------
# Dateien lesen
# ---------------------------------------------------------------------------

def read_sap_file(uploaded_file) -> Tuple[Dict[str, Set[int]], str, int]:
    """SAP wird bewusst nur über SAP Nummer + Liefertag gelesen.

    Keine Tournummer aus SAP. Keine Transportgruppe aus SAP.
    """
    excel = pd.ExcelFile(uploaded_file)
    sheet_name = excel.sheet_names[0]
    df = read_excel_with_detected_header(excel, sheet_name, kind="sap")
    if df.empty:
        return {}, sheet_name, 0

    columns = list(df.columns)
    sap_column = pick_column_by_name_or_position(
        columns,
        ["SAP", "SAP Nummer", "SAP-Nr", "SAP Nr", "Kundennummer", "Kunden Nummer"],
        SAP_COL_INDEX,
    )
    day_column = pick_column_by_name_or_position(
        columns,
        ["Liefertag", "Liefer Tag", "LT", "Tag", "Liefertag Code", "Liefertagcode"],
        SAP_DAY_COL_INDEX,
    )

    if sap_column is None or day_column is None:
        return {}, sheet_name, 0

    work = df[[sap_column, day_column]].copy()
    work.columns = ["sap", "tag"]
    work["sap"] = normalize_sap_series(work["sap"])
    work["tag_num"] = normalize_day_code_series(work["tag"])

    mask = (
        work["sap"].ne("")
        & work["tag_num"].between(1, 6, inclusive="both")
        & work["tag_num"].notna()
    )
    filtered = work.loc[mask, ["sap", "tag_num"]].copy()
    filtered["tag_int"] = filtered["tag_num"].astype(int)

    days_by_sap: Dict[str, Set[int]] = filtered.groupby("sap")["tag_int"].agg(set).to_dict()
    return days_by_sap, sheet_name, len(filtered)


def read_tourenplanung(uploaded_file) -> Tuple[pd.DataFrame, Dict[str, str], Dict[str, Dict[str, str]]]:
    """Liest nur den gewünschten Prüfbereich aus der Tourenplanung.

    NMS: komplett
    Malchow: komplett
    Direkt: nur Touren 1058, 2058, 3058, 4058, 5058, 6030
    """
    excel = pd.ExcelFile(uploaded_file)
    selected_sheets = select_relevant_sheets(excel)

    rows: List[dict] = []
    customer_info: Dict[str, Dict[str, str]] = {}

    for bereich, sheet_name in selected_sheets.items():
        df = read_excel_with_detected_header(excel, sheet_name, kind="tour")
        if df.empty:
            continue

        columns = list(df.columns)
        sap_column = pick_column_by_name_or_position(
            columns,
            ["SAP", "SAP Nummer", "SAP-Nr", "SAP Nr", "Kundennummer", "Kunden Nummer"],
            TOUR_SAP_COL_INDEX,
        )
        if sap_column is None:
            continue

        csb_column = pick_column_by_name_or_position(columns, ["CSB", "CSB Nummer", "CSB-Nr", "CSB Nr"], TOUR_CSB_COL_INDEX)
        name_column = pick_column_by_name_or_position(columns, ["Name", "Kundenname", "Marktname", "Kunde", "Bezeichnung", "Filialname"], 2)
        strasse_column = pick_column_by_name_or_position(columns, ["Strasse", "Straße", "Str", "Anschrift", "Adresse", "Strassenname", "Straßenname", "Strasse Hausnummer", "Straße Hausnummer"], 3)
        plz_column = pick_column_by_name_or_position(columns, ["Plz", "PLZ", "Postleitzahl"], 4)
        ort_column = pick_column_by_name_or_position(columns, ["Ort", "Stadt", "Plz Ort", "PLZ Ort", "Ortname"], 5)
        day_columns = get_tour_columns(df)

        work = df.copy()
        work["_sap"] = normalize_sap_series(work[sap_column])

        for row_index, row in work.iterrows():
            sap = value_to_clean_text(row.get("_sap", ""))
            if not sap:
                continue

            csb = value_to_clean_text(row.get(csb_column, "")) if csb_column else ""
            name = value_to_clean_text(row.get(name_column, "")) if name_column else ""
            strasse = value_to_clean_text(row.get(strasse_column, "")) if strasse_column else ""
            plz = value_to_clean_text(row.get(plz_column, "")) if plz_column else ""
            ort = value_to_clean_text(row.get(ort_column, "")) if ort_column else ""

            merge_info(customer_info, sap, {
                "csb": csb,
                "name": name,
                "strasse": strasse,
                "plz": plz,
                "ort": ort,
                "bereich": bereich,
                "blatt": sheet_name,
            })

            for day_num, day_column in day_columns.items():
                cell_value = row.get(day_column, "")
                if not day_value_is_set(cell_value):
                    continue

                all_tours = extract_tour_numbers(cell_value)

                if bereich == "Direkt":
                    selected_tours = [tour for tour in all_tours if tour in DIRECT_ALLOWED_TOURS]
                    for tour in selected_tours:
                        rows.append({
                            "Bereich": bereich,
                            "Blatt Tourenplanung": sheet_name,
                            "CSB": csb,
                            "SAP Nummer": sap,
                            "Name": name,
                            "Straße": strasse,
                            "PLZ": plz,
                            "Ort": ort,
                            "Liefertag Nr": day_num,
                            "Liefertag": f"{day_num} {DAY_NAMES[day_num]}",
                            "Tournummer Tour": tour,
                            "Quelle": "Tourenplanung",
                        })
                else:
                    # NMS und Malchow werden komplett geprüft.
                    tour_text = ", ".join(all_tours) if all_tours else value_to_clean_text(cell_value)
                    rows.append({
                        "Bereich": bereich,
                        "Blatt Tourenplanung": sheet_name,
                        "CSB": csb,
                        "SAP Nummer": sap,
                        "Name": name,
                        "Straße": strasse,
                        "PLZ": plz,
                        "Ort": ort,
                        "Liefertag Nr": day_num,
                        "Liefertag": f"{day_num} {DAY_NAMES[day_num]}",
                        "Tournummer Tour": tour_text,
                        "Quelle": "Tourenplanung",
                    })


    if not rows:
        empty = pd.DataFrame(columns=tour_scope_columns())
        return empty, selected_sheets, customer_info

    tour_df = pd.DataFrame(rows)
    tour_df = tour_df.drop_duplicates(
        subset=["Bereich", "Blatt Tourenplanung", "SAP Nummer", "Liefertag Nr", "Tournummer Tour"]
    ).reset_index(drop=True)
    tour_df = apply_final_scope_filter(tour_df)
    tour_df["_sap_sort"] = pd.to_numeric(tour_df["SAP Nummer"], errors="coerce").fillna(9_999_999_999)
    tour_df = tour_df.sort_values(["Bereich", "_sap_sort", "Liefertag Nr", "Tournummer Tour"]).drop(columns=["_sap_sort"]).reset_index(drop=True)
    return tour_df, selected_sheets, customer_info


# ---------------------------------------------------------------------------
# Vergleich
# ---------------------------------------------------------------------------

def tour_scope_columns() -> List[str]:
    return [
        "Bereich",
        "Blatt Tourenplanung",
        "CSB",
        "SAP Nummer",
        "Name",
        "Straße",
        "PLZ",
        "Ort",
        "Liefertag",
        "Tournummer Tour",
        "Quelle",
    ]


def result_columns() -> List[str]:
    return [
        "Prüfung",
        "Bereich",
        "Blatt Tourenplanung",
        "CSB",
        "SAP Nummer",
        "Name",
        "Straße",
        "PLZ",
        "Ort",
        "LT SAP",
        "LT Tourenplanung",
        "Gemeinsam",
        "Nur in Tour",
        "Nur in SAP",
        "Hinweis",
    ]


def aggregate_tour_days(tour_df: pd.DataFrame) -> Dict[str, Set[int]]:
    if tour_df.empty:
        return {}
    return tour_df.groupby("SAP Nummer")["Liefertag Nr"].agg(lambda x: set(int(v) for v in x)).to_dict()


def _tour_numbers_by_day(tour_df: pd.DataFrame) -> Dict[str, Dict[int, str]]:
    """Gibt pro SAP Nummer ein Dict {tag_nr: 'Tournummer(n)'} zurück."""
    if tour_df.empty:
        return {}
    out: Dict[str, Dict[int, str]] = {}
    for _, row in tour_df.iterrows():
        sap = str(row["SAP Nummer"])
        day = int(row["Liefertag Nr"])
        tn = value_to_clean_text(row.get("Tournummer Tour", ""))
        entry = out.setdefault(sap, {})
        existing = entry.get(day, "")
        if tn and tn not in existing:
            entry[day] = f"{existing}, {tn}" if existing else tn
    return out


def _short_day_list(days: Set[int]) -> str:
    return ", ".join(DAY_SHORT[d] for d in sorted(days))


def _day_list_with_tours(days: Set[int], tour_by_day: Dict[int, str]) -> str:
    """z.B. 'Do (44444), Sa (6001)'"""
    parts: List[str] = []
    for d in sorted(days):
        tn = tour_by_day.get(d, "")
        parts.append(f"{DAY_SHORT[d]} ({tn})" if tn else DAY_SHORT[d])
    return ", ".join(parts)


def _build_hinweis(sap_days: Set[int], tour_days: Set[int]) -> str:
    """Einzeiliger Hinweis pro Kunde."""
    gemeinsam = sap_days & tour_days
    nur_sap = sap_days - tour_days
    nur_tour = tour_days - sap_days

    if not nur_sap and not nur_tour:
        return "Kein Unterschied."
    if nur_sap and not nur_tour:
        return f"SAP enthält alle Tour-LT plus extra: {_short_day_list(nur_sap)}."
    if nur_tour and not nur_sap:
        return f"Tour enthält alle SAP-LT plus extra: {_short_day_list(nur_tour)}."
    if gemeinsam and tour_days <= sap_days:
        return f"SAP enthält alle Tour-LT plus extra: {_short_day_list(nur_sap)}."
    # Beide Seiten haben eigene
    return f"Abweichungen in beiden Richtungen."


def build_differences(
    tour_df: pd.DataFrame,
    days_by_sap: Dict[str, Set[int]],
    customer_info: Dict[str, Dict[str, str]],
) -> pd.DataFrame:
    """Ein Eintrag pro Kunde mit allen Abweichungen."""
    if tour_df.empty:
        return pd.DataFrame(columns=result_columns())

    tour_days_by_sap = aggregate_tour_days(tour_df)
    tn_by_day = _tour_numbers_by_day(tour_df)

    # Alle relevanten SAP-Nummern: aus Tour + aus SAP (soweit im Scope)
    all_saps = set(tour_df["SAP Nummer"].astype(str)) | set(days_by_sap.keys())
    # Nur SAPs die auch im Tourbereich vorkommen
    scope_saps = set(tour_df["SAP Nummer"].astype(str))
    all_saps = all_saps & scope_saps

    rows: List[dict] = []
    for sap in sorted(all_saps, key=lambda v: int(v) if str(v).isdigit() else 9_999_999_999):
        sap_days = days_by_sap.get(sap, set())
        tour_days = tour_days_by_sap.get(sap, set())

        nur_tour = tour_days - sap_days    # Fehlt in SAP
        nur_sap = sap_days - tour_days     # Fehlt in Tour
        gemeinsam = sap_days & tour_days

        if not nur_tour and not nur_sap:
            continue  # Kein Unterschied

        # Prüfung bestimmen
        if nur_tour and nur_sap:
            pruefung = "Beides"
        elif nur_tour:
            pruefung = "Fehlt in SAP"
        else:
            pruefung = "Fehlt in Tour"

        info = customer_info.get(sap, {})
        sap_tour_nums = tn_by_day.get(sap, {})

        # Kundeninfo aus Tour nehmen falls vorhanden
        first_tour_row = tour_df[tour_df["SAP Nummer"] == sap].iloc[0] if sap in tour_df["SAP Nummer"].values else None
        bereich = info.get("bereich", "")
        blatt = info.get("blatt", "")
        csb = info.get("csb", "")
        name = info.get("name", "")
        strasse = info.get("strasse", "")
        plz = info.get("plz", "")
        ort = info.get("ort", "")
        if first_tour_row is not None:
            bereich = bereich or value_to_clean_text(first_tour_row.get("Bereich", ""))
            blatt = blatt or value_to_clean_text(first_tour_row.get("Blatt Tourenplanung", ""))
            csb = csb or value_to_clean_text(first_tour_row.get("CSB", ""))
            name = name or value_to_clean_text(first_tour_row.get("Name", ""))

        rows.append({
            "Prüfung": pruefung,
            "Bereich": bereich,
            "Blatt Tourenplanung": blatt,
            "CSB": csb,
            "SAP Nummer": sap,
            "Name": name,
            "Straße": strasse,
            "PLZ": plz,
            "Ort": ort,
            "LT SAP": sorted_day_text(sap_days),
            "LT Tourenplanung": sorted_day_text(tour_days),
            "Gemeinsam": _short_day_list(gemeinsam) if gemeinsam else "–",
            "Nur in Tour": _day_list_with_tours(nur_tour, sap_tour_nums) if nur_tour else "–",
            "Nur in SAP": _short_day_list(nur_sap) if nur_sap else "–",
            "Hinweis": _build_hinweis(sap_days, tour_days),
        })

    return sort_result(pd.DataFrame(rows, columns=result_columns()))


def sort_result(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df if df is not None else pd.DataFrame(columns=result_columns())
    out = df.copy()
    bereich_order = {"NMS": 1, "Malchow": 2, "Direkt": 3}
    pruef_order = {"Beides": 1, "Fehlt in SAP": 2, "Fehlt in Tour": 3}
    out["_bereich_sort"] = out["Bereich"].map(bereich_order).fillna(99)
    out["_pruef_sort"] = out["Prüfung"].map(pruef_order).fillna(99)
    out["_sap_sort"] = pd.to_numeric(out["SAP Nummer"], errors="coerce").fillna(9_999_999_999)
    out = out.sort_values(["_bereich_sort", "_pruef_sort", "_sap_sort"]).drop(columns=["_bereich_sort", "_pruef_sort", "_sap_sort"])
    return out.reset_index(drop=True)


def build_direct_tourkunden(tour_df: pd.DataFrame) -> pd.DataFrame:
    if tour_df.empty:
        return pd.DataFrame(columns=tour_scope_columns())
    direct = tour_df[tour_df["Bereich"] == "Direkt"].copy()
    if direct.empty:
        return pd.DataFrame(columns=tour_scope_columns())
    # Harter Sicherheitsfilter: nur die sechs freigegebenen Touren
    direct["_tn_clean"] = direct["Tournummer Tour"].map(value_to_clean_text)
    direct = direct[direct["_tn_clean"].isin(DIRECT_ALLOWED_TOURS)].drop(columns=["_tn_clean"])
    if direct.empty:
        return pd.DataFrame(columns=tour_scope_columns())
    direct["_sap_sort"] = pd.to_numeric(direct["SAP Nummer"], errors="coerce").fillna(9_999_999_999)
    direct["_tag_sort"] = direct["Liefertag Nr"]
    direct = direct.sort_values(["_sap_sort", "_tag_sort", "Tournummer Tour"]).drop(columns=["_sap_sort", "_tag_sort"])
    return direct[tour_scope_columns()].reset_index(drop=True)


def build_direct_unique_customers(direct_tourkunden: pd.DataFrame) -> pd.DataFrame:
    cols = ["CSB", "SAP Nummer", "Name", "Straße", "PLZ", "Ort", "Gefundene Touren"]
    if direct_tourkunden.empty:
        return pd.DataFrame(columns=cols)
    agg = direct_tourkunden.groupby("SAP Nummer", as_index=False).agg(
        CSB=("CSB", "first"),
        Name=("Name", "first"),
        Straße=("Straße", "first"),
        PLZ=("PLZ", "first"),
        Ort=("Ort", "first"),
        Gefundene_Touren=("Tournummer Tour", lambda x: ", ".join(sorted(set(value_to_clean_text(v) for v in x if value_to_clean_text(v))))),
    )
    agg = agg.rename(columns={"Gefundene_Touren": "Gefundene Touren"})
    agg["_sap_sort"] = pd.to_numeric(agg["SAP Nummer"], errors="coerce").fillna(9_999_999_999)
    agg = agg.sort_values("_sap_sort").drop(columns=["_sap_sort"])
    return agg[cols].reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel-Ausgabe
# ---------------------------------------------------------------------------

_RIGHT_ALIGN_COLS = {"SAP Nummer", "CSB", "PLZ"}
_CENTER_ALIGN_COLS = {"Prüfung", "Bereich", "Quelle", "Liefertag Nr"}
_COL_WIDTH_HINTS = {
    "Prüfung": (14, 16),
    "Bereich": (10, 14),
    "Blatt Tourenplanung": (18, 28),
    "CSB": (10, 12),
    "SAP Nummer": (12, 14),
    "Name": (24, 44),
    "Straße": (20, 34),
    "PLZ": (8, 10),
    "Ort": (18, 30),
    "LT SAP": (24, 48),
    "LT Tourenplanung": (24, 48),
    "Gemeinsam": (16, 30),
    "Nur in Tour": (20, 48),
    "Nur in SAP": (16, 30),
    "Hinweis": (30, 58),
    "Quelle": (16, 24),
    "Liefertag": (16, 18),
    "Tournummer Tour": (16, 20),
    "Gefundene Touren": (18, 28),
}

# ---- Farben je Bereich ----
_BEREICH_FILL = {
    "NMS":     PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid"),   # blau-hell
    "Malchow": PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid"),   # grün-hell
    "Direkt":  PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid"),   # orange-hell
}
_BEREICH_FILL_ZEBRA = {
    "NMS":     PatternFill(start_color="FFC5D9F1", end_color="FFC5D9F1", fill_type="solid"),   # blau
    "Malchow": PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"),   # grün
    "Direkt":  PatternFill(start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid"),   # orange
}
_PRUEFUNG_FILL = {
    "Fehlt in SAP": PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid"),  # gelb
    "Fehlt in Tour": PatternFill(start_color="FFFCE4EC", end_color="FFFCE4EC", fill_type="solid"),  # rosa
    "Beides": PatternFill(start_color="FFFFE0B2", end_color="FFFFE0B2", fill_type="solid"),  # orange
}
_PRUEFUNG_FONT = {
    "Fehlt in SAP": Font(name="Calibri", size=10, bold=True, color="FF7F6000"),
    "Fehlt in Tour": Font(name="Calibri", size=10, bold=True, color="FFC00000"),
    "Beides": Font(name="Calibri", size=10, bold=True, color="FFE65100"),
}
_DEFAULT_ZEBRA = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")


def _format_sheet(writer, sheet_name: str, df: pd.DataFrame, use_bereich_colors: bool = True) -> None:
    ws = writer.sheets[sheet_name]
    n_rows = len(df)
    n_cols = len(df.columns)

    header_fill = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
    thin = Side(style="thin", color="FFCBD5E0")
    medium = Side(style="medium", color="FF305496")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    body_font = Font(name="Calibri", size=10)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=False)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    columns = list(df.columns)

    # --- Kopfzeile ---
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_thin
    ws.row_dimensions[1].height = 26

    # --- Datenzeilen ---
    bereich_values = df["Bereich"].tolist() if ("Bereich" in df.columns and use_bereich_colors) else []
    pruefung_values = df["Prüfung"].tolist() if "Prüfung" in df.columns else []

    for row_offset in range(n_rows):
        excel_row = row_offset + 2
        is_zebra = (row_offset % 2) == 1
        bereich = bereich_values[row_offset] if bereich_values else ""
        pruefung = pruefung_values[row_offset] if pruefung_values else ""
        new_group = bool(bereich_values) and row_offset > 0 and bereich_values[row_offset] != bereich_values[row_offset - 1]
        ws.row_dimensions[excel_row].height = 20

        # Zeile einfärben nach Bereich
        if use_bereich_colors and bereich in _BEREICH_FILL:
            row_fill = _BEREICH_FILL_ZEBRA[bereich] if is_zebra else _BEREICH_FILL[bereich]
        else:
            row_fill = _DEFAULT_ZEBRA if is_zebra else None

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.font = body_font
            if col_name in _CENTER_ALIGN_COLS:
                cell.alignment = align_center
            elif col_name in _RIGHT_ALIGN_COLS:
                cell.alignment = align_right
            else:
                cell.alignment = align_left
            if row_fill:
                cell.fill = row_fill
            top_side = medium if new_group else thin
            cell.border = Border(left=thin, right=thin, top=top_side, bottom=thin)

        # Prüfung-Spalte extra hervorheben
        if pruefung and "Prüfung" in columns:
            p_col = columns.index("Prüfung") + 1
            p_cell = ws.cell(row=excel_row, column=p_col)
            if pruefung in _PRUEFUNG_FILL:
                p_cell.fill = _PRUEFUNG_FILL[pruefung]
                p_cell.font = _PRUEFUNG_FONT[pruefung]

    # --- Spaltenbreiten ---
    for col_idx, col_name in enumerate(columns, start=1):
        sample = df[col_name].astype(str).head(300).tolist() if n_rows else []
        max_len = max([len(str(col_name))] + [len(v) for v in sample] + [8])
        min_w, max_w = _COL_WIDTH_HINTS.get(col_name, (12, 46))
        width = min(max(max_len + 2, min_w), max_w)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    if n_cols > 0:
        last_col = get_column_letter(n_cols)
        ws.auto_filter.ref = f"A1:{last_col}{n_rows + 1}"

    try:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.print_options.gridLines = False
        ws.print_title_rows = "1:1"
    except Exception:
        pass


def _write_summary_sheet(writer, counts: Dict[str, int], selected_sheets: Dict[str, str]) -> None:
    """Schreibt ein Zusammenfassungs-Blatt als erstes Blatt."""
    ws = writer.book.create_sheet("Zusammenfassung", 0)
    writer.sheets["Zusammenfassung"] = ws

    title_font = Font(name="Calibri", size=16, bold=True, color="FF305496")
    label_font = Font(name="Calibri", size=11, bold=True)
    value_font = Font(name="Calibri", size=11)
    small_font = Font(name="Calibri", size=9, color="FF666666")
    header_fill = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18

    ws.cell(row=2, column=2, value="SAP gegen Tourenplanung").font = title_font
    ws.cell(row=3, column=2, value="Nur geprüfter Tourbereich: NMS, Malchow, Direkt (6 Touren)").font = small_font

    row = 5
    ws.cell(row=row, column=2, value="Kennzahl").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=3, value="Wert").font = header_font
    ws.cell(row=row, column=3).fill = header_fill
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")

    items = [
        ("Geprüfte Tour-Zeilen", counts.get("tour_scope", 0)),
        ("  davon NMS", counts.get("nms", 0)),
        ("  davon Malchow", counts.get("malchow", 0)),
        ("  davon Direkt (Tourkunden)", counts.get("direkt", 0)),
        ("", ""),
        ("Kunden mit Abweichungen", counts.get("kunden_mit_diff", 0)),
        ("  davon Fehlt in SAP", counts.get("kunden_fehlt_sap", 0)),
        ("  davon Fehlt in Tour", counts.get("kunden_fehlt_tour", 0)),
        ("  davon Beides", counts.get("kunden_beides", 0)),
        ("", ""),
        ("Direkt: einmalige Kunden", counts.get("direct_unique", 0)),
    ]

    for i, (label, val) in enumerate(items, start=1):
        r = row + i
        ws.cell(row=r, column=2, value=label).font = label_font if label and not label.startswith(" ") else value_font
        if val != "":
            c = ws.cell(row=r, column=3, value=val)
            c.font = value_font
            c.alignment = Alignment(horizontal="right")

    # Geprüfte Blätter
    r = row + len(items) + 2
    ws.cell(row=r, column=2, value="Geprüfte Blätter").font = header_font
    ws.cell(row=r, column=2).fill = header_fill
    ws.cell(row=r, column=3, value="Blattname").font = header_font
    ws.cell(row=r, column=3).fill = header_fill
    for i, (bereich, blatt) in enumerate(selected_sheets.items(), start=1):
        ws.cell(row=r + i, column=2, value=bereich).font = label_font
        ws.cell(row=r + i, column=3, value=blatt).font = value_font

    # Direkt-Touren Info
    r2 = r + len(selected_sheets) + 2
    ws.cell(row=r2, column=2, value="Geprüfte Direkt-Touren").font = header_font
    ws.cell(row=r2, column=2).fill = header_fill
    ws.cell(row=r2, column=3).fill = header_fill
    for i, tour in enumerate(sorted(DIRECT_ALLOWED_TOURS), start=1):
        ws.cell(row=r2 + i, column=2, value=tour).font = value_font

    # Legende Hinweis-Spalte
    r3 = r2 + len(DIRECT_ALLOWED_TOURS) + 2
    ws.cell(row=r3, column=2, value="Legende Hinweis-Spalte").font = header_font
    ws.cell(row=r3, column=2).fill = header_fill
    ws.cell(row=r3, column=3).fill = header_fill
    ws.merge_cells(start_row=r3, start_column=2, end_row=r3, end_column=3)

    legende = [
        ("Gemeinsame LT", "Liefertage die in SAP und Tourenplanung übereinstimmen."),
        ("Nur in SAP", "Liefertage die nur in SAP stehen, nicht in der Tourenplanung."),
        ("Nur in Tour", "Liefertage die nur in der Tourenplanung stehen, nicht in SAP."),
        ("SAP enthält alle Tour-LT plus extra", "SAP hat alle Tage aus der Tour und zusätzlich weitere."),
        ("SAP enthält KEINE zusätzlichen LT", "SAP hat keine Tage über die Tourenplanung hinaus."),
        ("", ""),
        ("Kürzel", "Mo, Die, Mitt, Don, Fr, Sam"),
    ]
    for i, (bez, erkl) in enumerate(legende, start=1):
        ws.cell(row=r3 + i, column=2, value=bez).font = label_font if bez else value_font
        ws.cell(row=r3 + i, column=3, value=erkl).font = value_font

    ws.column_dimensions["C"].width = 56

    ws.sheet_properties.tabColor = "305496"


def build_excel(
    all_diff: pd.DataFrame,
    tour_scope: pd.DataFrame,
    direct_tourkunden: pd.DataFrame,
    direct_unique: pd.DataFrame,
    selected_sheets: Optional[Dict[str, str]] = None,
) -> bytes:
    # Gefilterte Ansichten
    missing_sap = all_diff[all_diff["Prüfung"].isin({"Fehlt in SAP", "Beides"})].copy() if not all_diff.empty else pd.DataFrame(columns=result_columns())
    missing_tour = all_diff[all_diff["Prüfung"].isin({"Fehlt in Tour", "Beides"})].copy() if not all_diff.empty else pd.DataFrame(columns=result_columns())

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheets = [
            ("Alle Unterschiede", all_diff, True),
            ("Fehlt in SAP", missing_sap, True),
            ("Fehlt in Tour", missing_tour, True),
            ("Geprüfter Tourbereich", tour_scope[tour_scope_columns()] if not tour_scope.empty else pd.DataFrame(columns=tour_scope_columns()), True),
            ("Direkt Tourkunden", direct_tourkunden, True),
            ("Direkt Kunden", direct_unique, False),
        ]
        for sheet_name, df, use_colors in sheets:
            df.to_excel(writer, index=False, sheet_name=sheet_name, na_rep="")
            _format_sheet(writer, sheet_name, df, use_bereich_colors=use_colors)

        # Zusammenfassung als erstes Blatt
        counts = {
            "tour_scope": len(tour_scope),
            "nms": int((tour_scope["Bereich"] == "NMS").sum()) if not tour_scope.empty and "Bereich" in tour_scope.columns else 0,
            "malchow": int((tour_scope["Bereich"] == "Malchow").sum()) if not tour_scope.empty and "Bereich" in tour_scope.columns else 0,
            "direkt": int((tour_scope["Bereich"] == "Direkt").sum()) if not tour_scope.empty and "Bereich" in tour_scope.columns else 0,
            "kunden_mit_diff": len(all_diff),
            "kunden_fehlt_sap": len(missing_sap),
            "kunden_fehlt_tour": len(missing_tour),
            "kunden_beides": int((all_diff["Prüfung"] == "Beides").sum()) if not all_diff.empty else 0,
            "direct_unique": len(direct_unique),
        }
        _write_summary_sheet(writer, counts, selected_sheets or {})

        # Tab-Farben setzen
        tab_colors = {
            "Alle Unterschiede": "C00000",
            "Fehlt in SAP": "ED7D31",
            "Fehlt in Tour": "FFC000",
            "Geprüfter Tourbereich": "305496",
            "Direkt Tourkunden": "548235",
            "Direkt Kunden": "7F7F7F",
        }
        wb = writer.book
        for ws in wb.worksheets:
            if ws.title in tab_colors:
                ws.sheet_properties.tabColor = tab_colors[ws.title]
            ws.sheet_state = "visible"
        if wb.worksheets:
            wb.active = 0
    return output.getvalue()

# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title="SAP gegen Tourenplanung", layout="wide")

st.title("SAP gegen Tourenplanung")
st.caption("Prüft NMS und Malchow komplett. Aus Direkt wird NICHT das ganze Blatt geprüft, sondern nur konkrete Tageszellen mit 1058, 2058, 3058, 4058, 5058 oder 6030.")

with st.container(border=True):
    st.markdown("**Prüflogik**")
    st.write(
        "SAP wird nur über **SAP Nummer + Liefertag** gelesen. "
        "Die Tournummer kommt ausschließlich aus der Tourenplanung. "
        "Der Rückvergleich wird nur für Kunden gemacht, die im geprüften Tourbereich vorkommen."
    )

col_a, col_b = st.columns(2)
with col_a:
    sap_datei = st.file_uploader(
        "SAP hochladen – SAP Nummer und Liefertag",
        type=["xlsx", "xlsm", "xls"],
        key="sap_datei",
    )
with col_b:
    tourenplanung_datei = st.file_uploader(
        "Tourenplanung hochladen – Blätter DIREKT, HUPA_NMS, HUPA_MALCHOW",
        type=["xlsx", "xlsm", "xls"],
        key="tourenplanung_datei",
    )

run = st.button("Excel erzeugen", type="primary", use_container_width=True)

if run:
    if not sap_datei or not tourenplanung_datei:
        st.error("Bitte SAP-Datei und Tourenplanung hochladen.")
        st.stop()

    try:
        days_by_sap, sap_sheet, sap_rows = read_sap_file(sap_datei)
        tour_scope, selected_sheets, customer_info = read_tourenplanung(tourenplanung_datei)

        selected_saps = set(tour_scope["SAP Nummer"].astype(str)) if not tour_scope.empty else set()
        # Wichtiger Filter: SAP-Rückrichtung nur für den tatsächlich geprüften Tourbereich.
        days_by_sap_scoped = {sap: days for sap, days in days_by_sap.items() if sap in selected_saps}

        all_diff = build_differences(tour_scope, days_by_sap_scoped, customer_info)

        direct_tourkunden = build_direct_tourkunden(tour_scope)
        direct_unique = build_direct_unique_customers(direct_tourkunden)

        # Harte Sicherheitsprüfung: Im Direktbereich dürfen nur die sechs freigegebenen Tournummern stehen.
        if not direct_tourkunden.empty and "Tournummer Tour" in direct_tourkunden.columns:
            tn_clean = direct_tourkunden["Tournummer Tour"].map(value_to_clean_text)
            bad_direct = direct_tourkunden[~tn_clean.isin(DIRECT_ALLOWED_TOURS)]
            if not bad_direct.empty:
                st.error("Direkt enthält ungefilterte Tournummern. Der Export wurde gestoppt, damit nicht das ganze Direkt-Blatt verglichen wird.")
                st.dataframe(bad_direct, use_container_width=True, hide_index=True)
                st.stop()

        excel_bytes = build_excel(all_diff, tour_scope, direct_tourkunden, direct_unique, selected_sheets)

        st.session_state["result"] = {
            "sap_sheet": sap_sheet,
            "sap_rows": sap_rows,
            "selected_sheets": selected_sheets,
            "tour_scope": tour_scope,
            "all_diff": all_diff,
            "direct_tourkunden": direct_tourkunden,
            "direct_unique": direct_unique,
            "excel_bytes": excel_bytes,
        }
    except Exception as exc:
        import traceback
        st.error(f"Fehler beim Verarbeiten der Dateien: {exc}")
        with st.expander("Technische Details", expanded=False):
            st.code(traceback.format_exc(), language="python")
        st.session_state.pop("result", None)

result = st.session_state.get("result")
if result:
    st.divider()
    left, right = st.columns([3, 1])
    with left:
        sheets_txt = ", ".join(f"{k}: {v}" for k, v in result["selected_sheets"].items())
        st.subheader("Ergebnis")
        st.caption(f"SAP Blatt: {result['sap_sheet']} · SAP Liefertage übernommen: {result['sap_rows']} · Tour-Blätter: {sheets_txt}")
    with right:
        st.download_button(
            "Excel herunterladen",
            data=result["excel_bytes"],
            file_name="sap_tourenplanung_gepruefter_tourbereich.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    diff = result["all_diff"]
    n_fehlt_sap = int((diff["Prüfung"].isin({"Fehlt in SAP", "Beides"})).sum()) if not diff.empty else 0
    n_fehlt_tour = int((diff["Prüfung"].isin({"Fehlt in Tour", "Beides"})).sum()) if not diff.empty else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Geprüfte Tour-Zeilen", len(result["tour_scope"]))
    c2.metric("Kunden mit Abweichungen", len(diff))
    c3.metric("Fehlt in SAP", n_fehlt_sap)
    c4.metric("Fehlt in Tour", n_fehlt_tour)

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Alle Unterschiede",
        "Fehlt in SAP",
        "Fehlt in Tour",
        "Geprüfter Tourbereich",
        "Direkt Tourkunden",
    ])

    with tab1:
        st.dataframe(diff, use_container_width=True, hide_index=True)
    with tab2:
        st.dataframe(diff[diff["Prüfung"].isin({"Fehlt in SAP", "Beides"})] if not diff.empty else diff, use_container_width=True, hide_index=True)
    with tab3:
        st.dataframe(diff[diff["Prüfung"].isin({"Fehlt in Tour", "Beides"})] if not diff.empty else diff, use_container_width=True, hide_index=True)
    with tab4:
        st.dataframe(result["tour_scope"][tour_scope_columns()] if not result["tour_scope"].empty else result["tour_scope"], use_container_width=True, hide_index=True)
    with tab5:
        st.markdown("**Einmalige Kunden auf den sechs Direkt-Touren**")
        st.dataframe(result["direct_unique"], use_container_width=True, hide_index=True)
        st.markdown("**Alle Direkt-Tour-Zeilen**")
        st.dataframe(result["direct_tourkunden"], use_container_width=True, hide_index=True)
